import urllib.request
import json
import io
from openpyxl import Workbook, load_workbook

auth_url = 'http://127.0.0.1:8000/api/v1/auth/login'
cat_url = 'http://127.0.0.1:8000/api/v1/admin/catalog/services?skip=0&limit=1000'

req = urllib.request.Request(
    auth_url,
    data=json.dumps({'email': 'admin@smartserve.com', 'password': 'AdminPassword123!'}).encode('utf-8'),
    headers={'Content-Type': 'application/json'}
)
token = json.loads(urllib.request.urlopen(req).read().decode())['access_token']
headers = {'Authorization': f'Bearer {token}', 'Content-Type': 'application/json'}

print("==================================================")
print("BULK OPERATIONS & IMPORT/EXPORT AUTOMATED AUDIT")
print("==================================================\n")

# STEP 1: Test Catalog Export & Parse XLSX Integrity
print("--- 1. TESTING CATALOG EXPORT & PARSING XLSX ---")
export_url = 'http://127.0.0.1:8000/api/v1/admin/catalog/export-excel'
req_export = urllib.request.Request(export_url, headers={'Authorization': f'Bearer {token}'})
xlsx_bytes = urllib.request.urlopen(req_export).read()

wb = load_workbook(filename=io.BytesIO(xlsx_bytes), data_only=True)
ws = wb.active

rows = list(ws.iter_rows(values_only=True))
print(f"  Export File Size: {len(xlsx_bytes)} bytes")
print(f"  Total Rows in Exported XLSX: {len(rows)} (1 Header + {len(rows)-1} Services)")

headers_row = list(rows[0])
headers_str = str(headers_row).replace('₹', 'Rs.')
print(f"  Headers: {headers_str}")

assert "Base Price (₹)" in headers_row, "Header 'Base Price (₹)' missing from export!"
assert len(rows) >= 399, f"Expected at least 399 service rows in export, got {len(rows)-1}"

first_data_row = rows[1]
assert isinstance(first_data_row[4], (int, float)), f"Expected numeric raw float for price, got {type(first_data_row[4])}"
print("  [OK] Excel export structure, headers, raw numeric price, and row counts verified.\n")

# STEP 2: Test Pre-Import Validation & Preview API
print("--- 2. TESTING PRE-IMPORT VALIDATION & PREVIEW API ---")
# Create test workbook
wb_test = Workbook()
ws_test = wb_test.active
ws_test.append(["Service ID", "Category", "Subcategory", "Service Name", "Base Price (₹)", "Max Demand Increase (%)", "Max Discount (%)", "Active"])

# Get an existing service ID
req_cat = urllib.request.Request(cat_url, headers={'Authorization': f'Bearer {token}'})
services = json.loads(urllib.request.urlopen(req_cat).read().decode())
pedicure = next(s for s in services if 'pedicure' in s['name'].lower())

ws_test.append([str(pedicure['id']), pedicure['category'], pedicure['subcategory'], pedicure['name'], 1199.0, 0.5, 0.3, "Yes"]) # Valid Update
ws_test.append(["", "Beauty", "Facials", "Test Temp Spa Service", 899.0, 0.2, 0.1, "Yes"]) # Valid Insert
ws_test.append(["", "Beauty", "Facials", "Broken Service", "INVALID_NUMERIC_PRICE", 0.2, 0.1, "Yes"]) # Invalid Price
ws_test.append([str(pedicure['id']), pedicure['category'], pedicure['subcategory'], pedicure['name'], 1199.0, 0.5, 0.3, "Yes"]) # Duplicate ID

test_stream = io.BytesIO()
wb_test.save(test_stream)
test_xlsx_bytes = test_stream.getvalue()

# Send Preview Request (Multipart Form)
boundary = '----WebKitFormBoundary7MA4YWxkTrZu0gW'
body = (
    f'--{boundary}\r\n'
    f'Content-Disposition: form-data; name="file"; filename="test_catalog.xlsx"\r\n'
    f'Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n'
).encode('utf-8') + test_xlsx_bytes + f'\r\n--{boundary}--\r\n'.encode('utf-8')

req_preview = urllib.request.Request(
    'http://127.0.0.1:8000/api/v1/admin/catalog/preview-import-excel',
    data=body,
    headers={
        'Authorization': f'Bearer {token}',
        'Content-Type': f'multipart/form-data; boundary={boundary}'
    }
)
preview_res = json.loads(urllib.request.urlopen(req_preview).read().decode())

print(f"  Preview Total Rows: {preview_res['total_rows']}")
print(f"  Preview Valid Count: {preview_res['valid_count']} (Updates: {preview_res['updates_count']}, New: {preview_res['new_count']})")
print(f"  Preview Invalid Count: {preview_res['invalid_count']}")
print(f"  Preview Errors Logged: {preview_res['errors']}")

assert preview_res['total_rows'] == 4, f"Expected 4 total rows in preview, got {preview_res['total_rows']}"
assert preview_res['invalid_count'] == 2, f"Expected 2 invalid rows (invalid price + duplicate ID), got {preview_res['invalid_count']}"
print("  [OK] Pre-Import Preview & Validation engine verified 100%.\n")

# STEP 3: Test Bulk Status Activate / Deactivate
print("--- 3. TESTING BULK STATUS ACTIVATE / DEACTIVATE ---")
svc1 = services[0]
svc2 = services[1]
test_ids = [svc1['id'], svc2['id']]

print(f"  Selected 2 Services for Bulk Status Test: '{svc1['name']}' and '{svc2['name']}'")

# Deactivate both
req_bulk_deact = urllib.request.Request(
    'http://127.0.0.1:8000/api/v1/admin/catalog/services/bulk-status',
    data=json.dumps({'service_ids': test_ids, 'is_active': False}).encode('utf-8'),
    headers=headers
)
deact_res = json.loads(urllib.request.urlopen(req_bulk_deact).read().decode())
print(f"  Bulk Deactivate Result: {deact_res['updated_count']} services updated to inactive")

# Verify in backend
services_fresh = json.loads(urllib.request.urlopen(req_cat).read().decode())
s1_fresh = next(s for s in services_fresh if s['id'] == svc1['id'])
s2_fresh = next(s for s in services_fresh if s['id'] == svc2['id'])

assert s1_fresh['is_active'] == False, f"Service '{svc1['name']}' failed to deactivate!"
assert s2_fresh['is_active'] == False, f"Service '{svc2['name']}' failed to deactivate!"
print("  [OK] Both selected services verified INACTIVE in database.")

# Reactivate both
req_bulk_act = urllib.request.Request(
    'http://127.0.0.1:8000/api/v1/admin/catalog/services/bulk-status',
    data=json.dumps({'service_ids': test_ids, 'is_active': True}).encode('utf-8'),
    headers=headers
)
act_res = json.loads(urllib.request.urlopen(req_bulk_act).read().decode())
print(f"  Bulk Reactivate Result: {act_res['updated_count']} services updated to active")

# Verify in backend
services_final = json.loads(urllib.request.urlopen(req_cat).read().decode())
s1_final = next(s for s in services_final if s['id'] == svc1['id'])
s2_final = next(s for s in services_final if s['id'] == svc2['id'])

assert s1_final['is_active'] == True, f"Service '{svc1['name']}' failed to reactivate!"
assert s2_final['is_active'] == True, f"Service '{svc2['name']}' failed to reactivate!"
print("  [OK] Both selected services restored to ACTIVE in database.\n")

# STEP 4: Final Catalog Integrity Verification
print("--- 4. VERIFYING DATABASE CATALOG INTEGRITY ---")
print(f"  Final Catalog Record Count in Neon Cloud PostgreSQL: {len(services_final)}")
assert len(services_final) >= 398, f"Database records corrupted! Count = {len(services_final)}"

print("\n==================================================")
print("BULK OPERATIONS & IMPORT/EXPORT PASSED 100%!")
print("==================================================")
