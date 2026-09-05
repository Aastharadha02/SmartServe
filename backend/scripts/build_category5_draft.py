import os
import sys
import json
import hashlib
import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Add parent path to import builder modules
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from category5_content_builder.carpenter_services_data import CARPENTER_SERVICES
from category5_content_builder.electrician_services_data import ELECTRICIAN_SERVICES
from category5_content_builder.plumber_services_data import PLUMBER_SERVICES

def build_category5_draft():
    all_builder_services = CARPENTER_SERVICES + ELECTRICIAN_SERVICES + PLUMBER_SERVICES
    print(f"Loaded {len(all_builder_services)} total services from builder modules.")
    assert len(all_builder_services) == 39, f"Expected 39 builder services, got {len(all_builder_services)}"
    
    # Load DB pre-change snapshot
    snapshot_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backups", "category5_electrician_plumber_carpenter_home_repairs_pre_change_snapshot.json"))
    with open(snapshot_path, "r", encoding="utf-8") as f:
        snapshot = json.load(f)
        
    db_services = {s["id"]: s for s in snapshot["services"]}
    print(f"Loaded {len(db_services)} database services from snapshot.")
    assert len(db_services) == 39, f"Expected 39 snapshot services, got {len(db_services)}"
    
    validated_draft_services = []
    
    for b_svc in all_builder_services:
        s_id = b_svc["id"]
        assert s_id in db_services, f"Builder service ID {s_id} ('{b_svc.get('name')}') not in DB snapshot!"
        db_s = db_services[s_id]
        
        # Verify exact identity parity
        assert b_svc["name"] == db_s["name"], f"Name mismatch on {s_id}: '{b_svc['name']}' vs '{db_s['name']}'"
        assert b_svc["category"] == db_s["category"], f"Category mismatch on {s_id}"
        assert b_svc["subcategory"] == db_s["subcategory"], f"Subcategory mismatch on {s_id}: '{b_svc['subcategory']}' vs '{db_s['subcategory']}'"
        assert b_svc["price"] == db_s["base_price"], f"Price mismatch on {s_id}: {b_svc['price']} vs {db_s['base_price']}"
        
        # Existing distinct features and real add-ons from DB
        distinct_features = list(db_s.get("distinct_features") or [])
        real_addons = db_s.get("real_addons") or []
        
        # Validate rich fields completeness
        for fld in ["description", "highlights", "included", "excluded", "process_steps", 
                    "tools_materials", "customer_setup", "aftercare", "expected_results", 
                    "important_notes", "warranty", "faqs", "dos", "donts", "tips"]:
            val = b_svc.get(fld)
            assert val is not None, f"Missing required field '{fld}' in service '{b_svc['name']}'"
            if isinstance(val, (list, str)):
                assert len(val) > 0, f"Empty field '{fld}' in service '{b_svc['name']}'"
                
        assert len(b_svc["highlights"]) >= 4, f"Highlights too short in '{b_svc['name']}'"
        assert len(b_svc["included"]) >= 4, f"Inclusions too short in '{b_svc['name']}'"
        assert len(b_svc["process_steps"]) >= 4, f"Process steps too short in '{b_svc['name']}'"
        assert len(b_svc["faqs"]) >= 4, f"FAQs count less than 4 in '{b_svc['name']}'"
        
        draft_record = {
            "id": s_id,
            "name": b_svc["name"],
            "category": b_svc["category"],
            "subcategory": b_svc["subcategory"],
            "price": b_svc["price"],
            "active": True,
            "description": b_svc["description"],
            "highlights": b_svc["highlights"],
            "included": b_svc["included"],
            "excluded": b_svc["excluded"],
            "process_steps": b_svc["process_steps"],
            "tools_materials": b_svc["tools_materials"],
            "customer_setup": b_svc["customer_setup"],
            "aftercare": b_svc["aftercare"],
            "expected_results": b_svc["expected_results"],
            "important_notes": b_svc["important_notes"],
            "warranty": b_svc["warranty"],
            "faqs": b_svc["faqs"],
            "dos": b_svc["dos"],
            "donts": b_svc["donts"],
            "tips": b_svc["tips"],
            "existing_add_ons": real_addons,
            "distinct_features_in_db": distinct_features
        }
        validated_draft_services.append(draft_record)
        
    print(f"[OK] All 39 Category 5 services validated against PostgreSQL with 100% parity.")
    
    # Sort logically by subcategory, then name
    validated_draft_services.sort(key=lambda s: (s["subcategory"], s["name"]))
    
    # Draft directory
    draft_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "catalog_drafts", "category5"))
    os.makedirs(draft_dir, exist_ok=True)
    
    # 1. Save JSON Draft
    json_path = os.path.join(draft_dir, "category5_electrician_plumber_carpenter_home_repairs_DRAFT.json")
    draft_doc = {
        "metadata": {
            "category": "5. Electrician, Plumber, Carpenter & Home Repairs",
            "draft_generated_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "total_services": len(validated_draft_services),
            "subcategories_count": 3,
            "subcategories": {
                "Carpenter": 17,
                "Electrician": 11,
                "Plumber": 11
            }
        },
        "services": validated_draft_services
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(draft_doc, f, indent=2, ensure_ascii=False)
        
    print(f"Saved Draft JSON: {json_path} ({os.path.getsize(json_path)} bytes)")
    
    # 2. Save XLSX Draft
    xlsx_path = os.path.join(draft_dir, "category5_electrician_plumber_carpenter_home_repairs_DRAFT.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Styles
    navy_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
    teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    blue_header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Sheet 1: CATEGORY_INDEX
    ws_index = wb.create_sheet(title="CATEGORY_INDEX")
    ws_index.views.sheetView[0].showGridLines = True
    
    ws_index.merge_cells("A1:C1")
    title_cell = ws_index["A1"]
    title_cell.value = "SmartServe Catalog - Category 5: Electrician, Plumber, Carpenter & Home Repairs Index"
    title_cell.font = title_font
    title_cell.fill = navy_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_index.row_dimensions[1].height = 30
    
    index_headers = ["subcategory", "service_count", "service_ids"]
    ws_index.append([])
    ws_index.append(index_headers)
    ws_index.row_dimensions[3].height = 22
    for col_idx in range(1, len(index_headers) + 1):
        cell = ws_index.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = teal_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        
    subcats = ["Carpenter", "Electrician", "Plumber"]
    for sc in subcats:
        svcs = [s for s in validated_draft_services if s["subcategory"] == sc]
        s_ids = ", ".join(s["id"] for s in svcs)
        ws_index.append([sc, len(svcs), s_ids])
        curr_row = ws_index.max_row
        for col_idx in range(1, 4):
            c = ws_index.cell(row=curr_row, column=col_idx)
            c.font = cell_font
            c.border = cell_border
            if col_idx == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
                
    ws_index.column_dimensions["A"].width = 24
    ws_index.column_dimensions["B"].width = 16
    ws_index.column_dimensions["C"].width = 85
    
    # Sheet 2: ALL_SERVICES
    ws_master = wb.create_sheet(title="ALL_SERVICES")
    ws_master.views.sheetView[0].showGridLines = True
    
    master_columns = [
        "service_id", "service_name", "category", "subcategory", "price", "active",
        "description", "highlights", "included", "excluded", "process_steps",
        "tools_materials", "customer_setup", "aftercare", "expected_results",
        "important_notes", "warranty", "faqs", "dos", "donts", "tips", "existing_add_ons"
    ]
    
    ws_master.append(master_columns)
    ws_master.row_dimensions[1].height = 25
    for col_idx in range(1, len(master_columns) + 1):
        cell = ws_master.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = navy_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = cell_border
        
    def format_service_row(s):
        hl_str = "\n".join(f"• {h}" for h in s["highlights"])
        inc_str = "\n".join(f"• {i}" for i in s["included"])
        exc_str = "\n".join(f"• {e}" for e in s["excluded"])
        proc_str = "\n\n".join(f"Step {p['step_number']}: {p['title']}\n{p['description']}" for p in s["process_steps"])
        tool_str = ", ".join(s["tools_materials"])
        setup_str = "\n".join(f"• {cs}" for cs in s["customer_setup"])
        after_str = "\n".join(f"• {ac}" for ac in s["aftercare"])
        exp_str = "\n".join(f"• {er}" for er in s["expected_results"])
        note_str = "\n".join(f"• {n}" for n in s["important_notes"])
        faq_str = "\n\n".join(f"Q: {f['question']}\nA: {f['answer']}" for f in s["faqs"])
        dos_str = "\n".join(f"• {d}" for d in s["dos"])
        dont_str = "\n".join(f"• {dt}" for dt in s["donts"])
        tips_str = "\n".join(f"• {t}" for t in s["tips"])
        addons_str = "\n".join(f"• {a.get('name')}: Rs.{a.get('price')}" for a in s["existing_add_ons"]) if s["existing_add_ons"] else "None"
        
        return [
            s["id"], s["name"], s["category"], s["subcategory"], s["price"], "True" if s["active"] else "False",
            s["description"], hl_str, inc_str, exc_str, proc_str,
            tool_str, setup_str, after_str, exp_str,
            note_str, s["warranty"] or "N/A", faq_str, dos_str, dont_str, tips_str, addons_str
        ]

    for s in validated_draft_services:
        row_data = format_service_row(s)
        ws_master.append(row_data)
        curr_row = ws_master.max_row
        ws_master.row_dimensions[curr_row].height = 65
        for col_idx in range(1, len(row_data) + 1):
            c = ws_master.cell(row=curr_row, column=col_idx)
            c.font = cell_font
            c.border = cell_border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            
    for col_idx in range(1, len(master_columns) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx in [1, 2, 4]:
            ws_master.column_dimensions[col_letter].width = 30
        elif col_idx in [3, 5, 6]:
            ws_master.column_dimensions[col_letter].width = 16
        else:
            ws_master.column_dimensions[col_letter].width = 45
            
    # Subcategory individual sheets: CARPENTER, ELECTRICIAN, PLUMBER
    for sc in subcats:
        sheet_name = sc.upper()
        ws_sub = wb.create_sheet(title=sheet_name)
        ws_sub.views.sheetView[0].showGridLines = True
        
        ws_sub.append(master_columns)
        ws_sub.row_dimensions[1].height = 25
        for col_idx in range(1, len(master_columns) + 1):
            cell = ws_sub.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = blue_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border
            
        sc_services = [s for s in validated_draft_services if s["subcategory"] == sc]
        for s in sc_services:
            row_data = format_service_row(s)
            ws_sub.append(row_data)
            curr_row = ws_sub.max_row
            ws_sub.row_dimensions[curr_row].height = 65
            for col_idx in range(1, len(row_data) + 1):
                c = ws_sub.cell(row=curr_row, column=col_idx)
                c.font = cell_font
                c.border = cell_border
                c.alignment = Alignment(vertical="top", wrap_text=True)
                
        for col_idx in range(1, len(master_columns) + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx in [1, 2, 4]:
                ws_sub.column_dimensions[col_letter].width = 30
            elif col_idx in [3, 5, 6]:
                ws_sub.column_dimensions[col_letter].width = 16
            else:
                ws_sub.column_dimensions[col_letter].width = 45
                
    wb.save(xlsx_path)
    print(f"Saved Draft XLSX: {xlsx_path} ({os.path.getsize(xlsx_path)} bytes)")
    
    # Calculate SHA256 of draft files
    with open(json_path, "rb") as f:
        json_sha = hashlib.sha256(f.read()).hexdigest()
    with open(xlsx_path, "rb") as f:
        xlsx_sha = hashlib.sha256(f.read()).hexdigest()
        
    print(f"Draft JSON SHA-256: {json_sha}")
    print(f"Draft XLSX SHA-256: {xlsx_sha}")
    
    return json_path, xlsx_path, json_sha, xlsx_sha, validated_draft_services

if __name__ == "__main__":
    build_category5_draft()
