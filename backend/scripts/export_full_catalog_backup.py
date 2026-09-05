"""
SmartServe Master Catalog Backup & Export Tool
Exports all 14 categories and 457 authentic services with complete, granular details
to Excel (.xlsx) and JSON (.json) formats in backend/backup/
"""

import os
import sys
import re
import json
import uuid
import datetime
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Ensure backend root is in sys.path
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from app.repositories.db import get_db
from app.models.service import Service

BACKUP_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backup"))
CATEGORIES_DIR = os.path.join(BACKUP_DIR, "categories")
os.makedirs(BACKUP_DIR, exist_ok=True)
os.makedirs(CATEGORIES_DIR, exist_ok=True)

def parse_cat_number(cat_name):
    m = re.match(r"^(\d+)\.", cat_name.strip())
    return int(m.group(1)) if m else 999

def format_bullet_list(items):
    if not items:
        return ""
    lines = []
    for it in items:
        if isinstance(it, dict):
            t = it.get("title") or it.get("name") or it.get("step")
            d = it.get("description") or it.get("desc") or it.get("details") or it.get("text")
            if t and d:
                lines.append(f"• {t}: {d}")
            elif d:
                lines.append(f"• {d}")
            elif t:
                lines.append(f"• {t}")
            else:
                lines.append(f"• {json.dumps(it)}")
        elif isinstance(it, str):
            lines.append(f"• {it}")
        else:
            lines.append(f"• {str(it)}")
    return "\n".join(lines)

def format_process_steps(steps):
    if not steps:
        return ""
    lines = []
    for idx, st in enumerate(steps, 1):
        if isinstance(st, dict):
            num = st.get("step_number") or idx
            title = st.get("title") or f"Step {num}"
            desc = st.get("description") or ""
            dur = st.get("duration_minutes")
            dur_str = f" ({dur} mins)" if dur else ""
            lines.append(f"{num}. {title}{dur_str}: {desc}")
        elif isinstance(st, str):
            lines.append(f"{idx}. {st}")
        else:
            lines.append(f"{idx}. {str(st)}")
    return "\n".join(lines)

def format_faqs(faqs):
    if not faqs:
        return ""
    lines = []
    for idx, f in enumerate(faqs, 1):
        if isinstance(f, dict):
            q = f.get("question") or f.get("q") or ""
            a = f.get("answer") or f.get("a") or ""
            lines.append(f"Q{idx}: {q}\nA{idx}: {a}\n")
        elif isinstance(f, str):
            lines.append(f"• {f}\n")
    return "\n".join(lines).strip()

def format_addons(addons):
    if not addons:
        return ""
    lines = []
    for a in addons:
        if isinstance(a, dict):
            name = a.get("name") or "Add-on"
            price = a.get("price", 0)
            desc = a.get("description") or ""
            desc_str = f" - {desc}" if desc else ""
            lines.append(f"• {name} (₹{price:,.0f}){desc_str}")
        elif isinstance(a, str):
            lines.append(f"• {a}")
    return "\n".join(lines)

def extract_service_full_details(s):
    df = s.distinct_features if isinstance(s.distinct_features, (dict, list)) else {}
    sa = s.suggested_addons if isinstance(s.suggested_addons, list) else []

    description = ""
    highlights = []
    included = []
    excluded = []
    warranty = ""
    faqs = []
    process_steps = []
    tools_materials = []
    customer_setup = []
    aftercare = []
    expected_results = []
    important_notes = []
    tips = []
    dos = []
    donts = []
    duration = 60
    addons = []
    seo_title = ""
    seo_description = ""
    keywords = []

    if isinstance(df, dict):
        description = df.get("description", "")
        highlights = list(df.get("highlights", []))
        included = list(df.get("included", []))
        excluded = list(df.get("excluded", []))
        warranty = df.get("warranty", "")
        faqs = list(df.get("faqs", []))
    elif isinstance(df, list):
        included = [str(x) for x in df]

    for item in sa:
        if not isinstance(item, dict):
            continue
        itype = item.get("type")
        if itype in ["description", "service_description"] and not description:
            description = item.get("text") or item.get("description") or ""
        elif itype == "highlights" and not highlights:
            highlights = item.get("highlights") or item.get("items") or []
        elif itype in ["excluded", "excluded_scope"] and not excluded:
            excluded = item.get("excluded") or item.get("items") or []
        elif itype == "process_steps":
            process_steps = item.get("steps") or item.get("items") or []
        elif itype in ["tools_materials", "products_and_tools"]:
            tools = item.get("tools") or []
            mats = item.get("materials") or []
            pt = item.get("products_and_tools") or item.get("items") or []
            tools_materials = pt if pt else tools + mats
        elif itype in ["customer_setup", "preparation"]:
            customer_setup = item.get("setup") or item.get("requirements") or item.get("items") or []
        elif itype == "aftercare_precautions":
            aftercare = item.get("aftercare") or item.get("items") or []
        elif itype == "expected_results":
            res = item.get("items") or item.get("results") or []
            expected_results = res if isinstance(res, list) else [str(res)]
        elif itype == "important_notes":
            notes = item.get("items") or item.get("notes") or []
            important_notes = notes if isinstance(notes, list) else [str(notes)]
        elif itype == "warranty" and not warranty:
            warranty = item.get("details") or item.get("warranty") or ""
        elif itype == "faqs" and not faqs:
            faqs = item.get("items") or item.get("faqs") or []
        elif itype == "tips":
            tips = item.get("items") or item.get("tips") or []
        elif itype in ["dos_donts", "dos_and_donts"]:
            dos = item.get("dos") or []
            donts = item.get("donts") or []
        elif itype in ["duration", "estimated_duration"]:
            mins = item.get("minutes") or item.get("duration") or item.get("duration_minutes")
            if mins:
                try:
                    duration = int(mins)
                except Exception:
                    pass
        elif itype == "seo_metadata":
            seo_title = item.get("seo_title", "")
            seo_description = item.get("seo_description", "")
            keywords = item.get("keywords", [])
        elif "price" in item and ("name" in item or "addon_id" in item):
            addons.append(item)

    if not description:
        description = f"Professional {s.name} service delivered by verified specialists under {s.category}."

    if not warranty:
        warranty = "SmartServe 100% Quality & Service Satisfaction Assurance."

    cat_num = parse_cat_number(s.category)

    return {
        "service_id": str(s.id),
        "category_number": cat_num,
        "category_name": s.category,
        "subcategory_name": s.subcategory,
        "service_name": s.name,
        "base_price_inr": float(s.base_price),
        "estimated_duration_minutes": duration,
        "max_demand_surge_percent": float(s.max_demand_increase or 0.0),
        "max_discount_percent": float(s.max_discount or 0.0),
        "is_active": bool(s.is_active),
        "description": description,
        "highlights": highlights,
        "included_features": included,
        "excluded_scope": excluded,
        "process_steps": process_steps,
        "tools_materials": tools_materials,
        "customer_setup_requirements": customer_setup,
        "aftercare_precautions": aftercare,
        "expected_results": expected_results,
        "important_notes": important_notes,
        "warranty_coverage": warranty,
        "faqs": faqs,
        "tips": tips,
        "dos": dos,
        "donts": donts,
        "suggested_addons": addons,
        "seo_title": seo_title or f"{s.name} | SmartServe",
        "seo_description": seo_description or description,
        "keywords": keywords,
        "created_at": s.created_at.isoformat() if hasattr(s.created_at, "isoformat") else str(s.created_at),
        "updated_at": s.updated_at.isoformat() if hasattr(s.updated_at, "isoformat") else str(s.updated_at),
    }

def apply_excel_styling(ws, header_fill_color="1E3A8A"):
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=header_fill_color, end_color=header_fill_color, fill_type="solid")
    regular_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="top", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="top")
    thin_border = Border(
        left=Side(style="thin", color="CBD5E1"),
        right=Side(style="thin", color="CBD5E1"),
        top=Side(style="thin", color="CBD5E1"),
        bottom=Side(style="thin", color="CBD5E1")
    )

    ws.row_dimensions[1].height = 28
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = regular_font
            cell.border = thin_border
            header_name = str(ws.cell(row=1, column=col).value).lower()
            if any(k in header_name for k in ["price", "duration", "surge", "discount", "number", "count", "active"]):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    ws.freeze_panes = "A2"

def export_all():
    print("=" * 70)
    print("SmartServe Master Catalog Backup & Granular Data Exporter")
    print("=" * 70)

    db = next(get_db())
    raw_services = db.query(Service).all()
    print(f"Loaded {len(raw_services)} total services from PostgreSQL database.")

    # Process all services
    processed = [extract_service_full_details(s) for s in raw_services]

    # Sort by Category Number, Subcategory, Name
    processed.sort(key=lambda x: (x["category_number"], x["category_name"], x["subcategory_name"], x["service_name"]))

    # Groupings
    categories_dict = {}
    for svc in processed:
        cat_name = svc["category_name"]
        subcat_name = svc["subcategory_name"]
        if cat_name not in categories_dict:
            categories_dict[cat_name] = {
                "category_number": svc["category_number"],
                "category_name": cat_name,
                "subcategories": {},
                "services": []
            }
        categories_dict[cat_name]["services"].append(svc)
        if subcat_name not in categories_dict[cat_name]["subcategories"]:
            categories_dict[cat_name]["subcategories"][subcat_name] = []
        categories_dict[cat_name]["subcategories"][subcat_name].append(svc)

    print(f"Organized into {len(categories_dict)} categories.")

    # 1. GENERATE MASTER JSON
    print("\n--> [1/4] Generating Master JSON Backups...")
    timestamp_str = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    master_json_data = {
        "metadata": {
            "app": "SmartServe",
            "environment": "Production / Staging / Local",
            "backup_timestamp": timestamp_str,
            "total_categories": len(categories_dict),
            "total_services": len(processed),
            "description": "Complete master catalog backup with full granular service details, pricing, process workflows, checklists, and FAQs."
        },
        "catalog_summary": [
            {
                "category_number": cdata["category_number"],
                "category_name": cdata["category_name"],
                "subcategory_count": len(cdata["subcategories"]),
                "service_count": len(cdata["services"]),
                "min_price_inr": min(s["base_price_inr"] for s in cdata["services"]),
                "max_price_inr": max(s["base_price_inr"] for s in cdata["services"]),
                "avg_price_inr": round(sum(s["base_price_inr"] for s in cdata["services"]) / len(cdata["services"]), 2),
                "subcategories": list(cdata["subcategories"].keys())
            }
            for cdata in categories_dict.values()
        ],
        "all_services": processed
    }

    master_json_path = os.path.join(BACKUP_DIR, "SmartServe_Master_Catalog_All_Services.json")
    with open(master_json_path, "w", encoding="utf-8") as f:
        json.dump(master_json_data, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {master_json_path} ({os.path.getsize(master_json_path):,} bytes)")

    # Standard filename alias
    alias_json_path = os.path.join(BACKUP_DIR, "smartserve_complete_catalog_backup.json")
    with open(alias_json_path, "w", encoding="utf-8") as f:
        json.dump(master_json_data, f, indent=2, ensure_ascii=False)
    print(f"    Saved: {alias_json_path} ({os.path.getsize(alias_json_path):,} bytes)")

    # 2. GENERATE MASTER EXCEL WORKBOOK (.XLSX)
    print("\n--> [2/4] Generating Master Excel Workbook (.xlsx)...")
    wb = openpyxl.Workbook()

    columns_def = [
        ("Cat #", 8),
        ("Category Name", 28),
        ("Subcategory Name", 28),
        ("Service Name", 32),
        ("Price (₹)", 12),
        ("Duration (Mins)", 14),
        ("Description", 45),
        ("Highlights", 40),
        ("What Is Included", 40),
        ("What Is Excluded", 38),
        ("Process Steps", 45),
        ("Tools & Materials", 35),
        ("Customer Setup & Prerequisites", 35),
        ("Aftercare Guidance", 35),
        ("Expected Results", 35),
        ("Important Notes", 35),
        ("Warranty Coverage", 30),
        ("Frequently Asked Questions", 45),
        ("Professional Tips", 35),
        ("Do's", 30),
        ("Don'ts", 30),
        ("Suggested Add-Ons", 32),
        ("Max Surge %", 12),
        ("Max Discount %", 14),
        ("Is Active", 10),
        ("Service UUID", 38),
        ("Created At", 22),
        ("Updated At", 22)
    ]

    def build_service_row(s):
        return [
            s["category_number"],
            s["category_name"],
            s["subcategory_name"],
            s["service_name"],
            s["base_price_inr"],
            s["estimated_duration_minutes"],
            s["description"],
            format_bullet_list(s["highlights"]),
            format_bullet_list(s["included_features"]),
            format_bullet_list(s["excluded_scope"]),
            format_process_steps(s["process_steps"]),
            format_bullet_list(s["tools_materials"]),
            format_bullet_list(s["customer_setup_requirements"]),
            format_bullet_list(s["aftercare_precautions"]),
            format_bullet_list(s["expected_results"]),
            format_bullet_list(s["important_notes"]),
            s["warranty_coverage"],
            format_faqs(s["faqs"]),
            format_bullet_list(s["tips"]),
            format_bullet_list(s["dos"]),
            format_bullet_list(s["donts"]),
            format_addons(s["suggested_addons"]),
            s["max_demand_surge_percent"],
            s["max_discount_percent"],
            "Yes" if s["is_active"] else "No",
            s["service_id"],
            s["created_at"],
            s["updated_at"]
        ]

    # Sheet 1: Master Sheet (All Services)
    ws_all = wb.active
    ws_all.title = "Master_All_Services"
    ws_all.views.sheetView[0].showGridLines = True
    ws_all.append([col[0] for col in columns_def])

    for s in processed:
        ws_all.append(build_service_row(s))

    for idx, col in enumerate(columns_def, 1):
        col_letter = get_column_letter(idx)
        ws_all.column_dimensions[col_letter].width = col[1]

    apply_excel_styling(ws_all, header_fill_color="0F291E")

    # Sheet 2: Categories Overview
    ws_cats = wb.create_sheet(title="Categories_Overview")
    ws_cats.views.sheetView[0].showGridLines = True
    cat_headers = [
        ("Cat #", 8),
        ("Category Name", 35),
        ("Subcategories Count", 18),
        ("Total Services Count", 18),
        ("Min Price (₹)", 14),
        ("Max Price (₹)", 14),
        ("Avg Price (₹)", 14),
        ("Subcategories Covered", 60)
    ]
    ws_cats.append([c[0] for c in cat_headers])
    for cdata in categories_dict.values():
        prices = [s["base_price_inr"] for s in cdata["services"]]
        ws_cats.append([
            cdata["category_number"],
            cdata["category_name"],
            len(cdata["subcategories"]),
            len(cdata["services"]),
            min(prices),
            max(prices),
            round(sum(prices) / len(prices), 2),
            ", ".join(cdata["subcategories"].keys())
        ])
    for idx, col in enumerate(cat_headers, 1):
        ws_cats.column_dimensions[get_column_letter(idx)].width = col[1]
    apply_excel_styling(ws_cats, header_fill_color="1E3A8A")

    # Sheets 3 to 16: Category-Specific Sheets
    for cat_name, cdata in categories_dict.items():
        # Shorten sheet title to 31 chars max
        short_title = f"{cdata['category_number']}. {cat_name.split('.', 1)[-1].strip()}"
        if len(short_title) > 31:
            short_title = short_title[:31]
        
        ws_cat = wb.create_sheet(title=short_title)
        ws_cat.views.sheetView[0].showGridLines = True
        ws_cat.append([col[0] for col in columns_def])

        for s in cdata["services"]:
            ws_cat.append(build_service_row(s))

        for idx, col in enumerate(columns_def, 1):
            ws_cat.column_dimensions[get_column_letter(idx)].width = col[1]

        apply_excel_styling(ws_cat, header_fill_color="1E40AF")

    master_xlsx_path = os.path.join(BACKUP_DIR, "SmartServe_Master_Catalog_All_Services.xlsx")
    wb.save(master_xlsx_path)
    print(f"    Saved: {master_xlsx_path} ({os.path.getsize(master_xlsx_path):,} bytes)")

    alias_xlsx_path = os.path.join(BACKUP_DIR, "smartserve_complete_catalog_backup.xlsx")
    wb.save(alias_xlsx_path)
    print(f"    Saved: {alias_xlsx_path} ({os.path.getsize(alias_xlsx_path):,} bytes)")

    # 3. GENERATE CATEGORY-SPECIFIC INDIVIDUAL JSON & XLSX FILES
    print("\n--> [3/4] Generating Individual Category Backups in backend/backup/categories/...")
    for cat_name, cdata in categories_dict.items():
        cat_num = cdata["category_number"]
        clean_slug = re.sub(r"[^\w\-_]", "_", cat_name.lower().replace(" & ", "_").replace(" ", "_"))
        clean_slug = re.sub(r"_+", "_", clean_slug).strip("_")

        # Category JSON
        cat_json_path = os.path.join(CATEGORIES_DIR, f"Category_{cat_num:02d}_{clean_slug}.json")
        with open(cat_json_path, "w", encoding="utf-8") as f:
            json.dump({
                "category_number": cat_num,
                "category_name": cat_name,
                "service_count": len(cdata["services"]),
                "subcategories": list(cdata["subcategories"].keys()),
                "services": cdata["services"]
            }, f, indent=2, ensure_ascii=False)

        # Category XLSX
        cat_wb = openpyxl.Workbook()
        cat_ws = cat_wb.active
        cat_ws.title = f"Category_{cat_num}"
        cat_ws.views.sheetView[0].showGridLines = True
        cat_ws.append([col[0] for col in columns_def])
        for s in cdata["services"]:
            cat_ws.append(build_service_row(s))
        for idx, col in enumerate(columns_def, 1):
            cat_ws.column_dimensions[get_column_letter(idx)].width = col[1]
        apply_excel_styling(cat_ws, header_fill_color="1E3A8A")

        cat_xlsx_path = os.path.join(CATEGORIES_DIR, f"Category_{cat_num:02d}_{clean_slug}.xlsx")
        cat_wb.save(cat_xlsx_path)
        print(f"    Saved: Category {cat_num:02d} ({len(cdata['services'])} services) -> .xlsx & .json")

    # 4. GENERATE DOCUMENTATION / README
    print("\n--> [4/4] Writing README.md in backend/backup/...")
    readme_path = os.path.join(BACKUP_DIR, "README.md")
    with open(readme_path, "w", encoding="utf-8") as f:
        f.write(f"""# SmartServe Master Service Catalog Backup & Data Export

**Generated On**: {timestamp_str}  
**Total Categories**: {len(categories_dict)}  
**Total Verified Services**: {len(processed)}  
**Target Recipient**: Aastha & Stakeholders  

---

## 📁 Backup Files Summary

| File Name | Format | Description |
| :--- | :--- | :--- |
| `SmartServe_Master_Catalog_All_Services.xlsx` | **Excel (.xlsx)** | Master workbook containing all 457 services in a single sheet + 1 summary sheet + 14 category-specific tabs. |
| `SmartServe_Master_Catalog_All_Services.json` | **JSON (.json)** | Full JSON export with metadata, summaries, and nested granular attributes for all 457 services. |
| `smartserve_complete_catalog_backup.xlsx` | **Excel (.xlsx)** | Direct alias of the master Excel file for quick reference. |
| `smartserve_complete_catalog_backup.json` | **JSON (.json)** | Direct alias of the master JSON file for quick reference. |
| `categories/Category_XX_*.xlsx` | **Excel (.xlsx)** | 14 standalone category workbooks for focused category reviews. |
| `categories/Category_XX_*.json` | **JSON (.json)** | 14 standalone category JSON payloads for API/data pipelines. |

---

## 📋 Comprehensive Fields Captured for Every Service

Each service in both the Excel and JSON files contains the complete operational, marketing, and execution specifications:
1. **Category Number & Name**: Standardized 1-14 hierarchy.
2. **Subcategory Name**: The exact domain classification.
3. **Service Name**: Realistic, authentic industry title (0 generic placeholders).
4. **Base Price (₹)**: Current standard pricing in INR.
5. **Estimated Duration (Minutes)**: Realistic completion time.
6. **Description**: Clear, customer-facing service overview.
7. **Highlights**: Key differentiators and selling propositions.
8. **What Is Included**: Exact bulleted procedure and standard deliverables.
9. **What Is Excluded**: Clearly defined boundary and out-of-scope items.
10. **Process Steps**: Numbered sequence with title, description, and time.
11. **Tools & Materials**: Equipment, safety supplies, and products used.
12. **Customer Setup & Prerequisites**: What the client needs to prepare in advance.
13. **Aftercare Guidance**: Post-service precautions and maintenance advice.
14. **Expected Results**: Definitive outcome and quality expectations.
15. **Important Notes**: Booking conditions, cancellation, and safety caveats.
16. **Warranty Coverage**: Explicit satisfaction and re-service guarantees.
17. **Frequently Asked Questions (FAQs)**: Realistic customer Q&As.
18. **Professional Tips**: Expert guidance for extended longevity.
19. **Do's & Don'ts**: Best practices and risks to avoid.
20. **Suggested Add-Ons**: Complimentary packages with add-on pricing.
21. **Pricing Flexibility**: Maximum surge demand % and maximum discount %.
22. **Database Identifiers**: UUID, created timestamp, and last updated timestamp.

---

## 📊 Catalog Category Breakdown

| # | Category | Services Count | Subcategories Count | Price Range (₹) | Avg Price (₹) |
| :---: | :--- | :---: | :---: | :---: | :---: |
""")
        for cdata in categories_dict.values():
            prices = [s["base_price_inr"] for s in cdata["services"]]
            f.write(f"| {cdata['category_number']} | {cdata['category_name']} | {len(cdata['services'])} | {len(cdata['subcategories'])} | ₹{min(prices):,.0f} - ₹{max(prices):,.0f} | ₹{sum(prices)/len(prices):,.2f} |\n")
        f.write(f"\n**TOTAL**: {len(processed)} Services across {len(categories_dict)} Categories.\n")

    print(f"    Saved: {readme_path}")
    print("\n[COMPLETE] All catalog backup files generated successfully in backend/backup/!")

if __name__ == "__main__":
    export_all()
