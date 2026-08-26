import io
import uuid
from typing import List, Dict, Any, Tuple, Optional
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.models.service import Service
from app.repositories import service_repository, audit_repository

def generate_catalog_excel(
    db: Session, 
    category: Optional[str] = None, 
    subcategory: Optional[str] = None, 
    search: Optional[str] = None
) -> bytes:
    """Generate Excel binary (.xlsx) containing catalog services with optional filters."""
    services = service_repository.get_services(db, limit=1000)
    
    if category:
        services = [s for s in services if s.category.lower() == category.lower()]
    if subcategory:
        services = [s for s in services if s.subcategory.lower() == subcategory.lower()]
    if search:
        s_lower = search.lower()
        services = [s for s in services if s_lower in s.name.lower() or s_lower in s.category.lower() or s_lower in s.subcategory.lower()]

    wb = Workbook()
    ws = wb.active
    ws.title = "Service Catalog"

    headers = [
        "Service ID", "Category", "Subcategory", "Service Name",
        "Base Price (₹)", "Max Demand Increase (%)", "Max Discount (%)", "Active"
    ]
    ws.append(headers)

    for s in services:
        ws.append([
            str(s.id),
            s.category,
            s.subcategory,
            s.name,
            s.base_price,
            s.max_demand_increase,
            s.max_discount,
            "Yes" if s.is_active else "No"
        ])

    stream = io.BytesIO()
    wb.save(stream)
    return stream.getvalue()

def preview_import_catalog_excel(db: Session, file_contents: bytes) -> Dict[str, Any]:
    """Parse uploaded Excel spreadsheet and perform pre-import validation preview without modifying database."""
    wb = load_workbook(filename=io.BytesIO(file_contents), data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return {
            "total_rows": 0,
            "valid_count": 0,
            "updates_count": 0,
            "new_count": 0,
            "invalid_count": 1,
            "errors": ["Excel spreadsheet is empty or missing header row."],
            "rows_preview": []
        }

    rows_preview = []
    errors = []
    updates_count = 0
    new_count = 0
    invalid_count = 0

    seen_service_ids = set()

    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue

        raw_id = str(row[0]).strip() if len(row) > 0 and row[0] is not None else ""
        raw_cat = str(row[1]).strip() if len(row) > 1 and row[1] is not None else "General"
        raw_sub = str(row[2]).strip() if len(row) > 2 and row[2] is not None else "Standard"
        raw_name = str(row[3]).strip() if len(row) > 3 and row[3] is not None else "New Service"
        raw_price = row[4] if len(row) > 4 else None
        raw_surge = row[5] if len(row) > 5 else 0.0
        raw_discount = row[6] if len(row) > 6 else 0.0

        row_errors = []

        # Validate Price
        base_price = 0.0
        if raw_price is None or str(raw_price).strip() == "":
            row_errors.append(f"Row {idx}: Invalid Base Price — expected numeric value.")
        else:
            try:
                base_price = float(raw_price)
                if base_price < 0:
                    row_errors.append(f"Row {idx}: Base price cannot be negative.")
            except (ValueError, TypeError):
                row_errors.append(f"Row {idx}: Invalid Base Price '{raw_price}' — expected numeric value.")

        # Check Duplicate IDs
        if raw_id:
            if raw_id in seen_service_ids:
                row_errors.append(f"Row {idx}: Duplicate Service ID '{raw_id}' detected in spreadsheet.")
            else:
                seen_service_ids.add(raw_id)

        # Check ID in DB
        action_type = "INSERT"
        existing_service = None
        if raw_id:
            try:
                s_uuid = uuid.UUID(raw_id)
                existing_service = service_repository.get_service_by_id(db, s_uuid)
                if existing_service:
                    action_type = "UPDATE"
                else:
                    action_type = "INSERT (New ID)"
            except ValueError:
                row_errors.append(f"Row {idx}: Invalid Service ID UUID format '{raw_id}'.")

        if row_errors:
            invalid_count += 1
            errors.extend(row_errors)
            rows_preview.append({
                "row_number": idx,
                "service_id": raw_id,
                "name": raw_name,
                "category": raw_cat,
                "subcategory": raw_sub,
                "base_price": str(raw_price),
                "action_type": action_type,
                "status": "INVALID",
                "error": "; ".join(row_errors)
            })
        else:
            if action_type == "UPDATE":
                updates_count += 1
            else:
                new_count += 1

            rows_preview.append({
                "row_number": idx,
                "service_id": str(existing_service.id) if existing_service else raw_id,
                "name": raw_name,
                "category": raw_cat,
                "subcategory": raw_sub,
                "base_price": base_price,
                "action_type": action_type,
                "status": "VALID",
                "error": None
            })

    total_valid = updates_count + new_count

    return {
        "total_rows": len(rows_preview),
        "valid_count": total_valid,
        "updates_count": updates_count,
        "new_count": new_count,
        "invalid_count": invalid_count,
        "errors": errors,
        "rows_preview": rows_preview
    }

def parse_and_import_catalog_excel(db: Session, file_contents: bytes, actor_email: str) -> Tuple[int, int, List[str]]:
    """Parse uploaded Excel catalog (.xlsx) and bulk insert/update service catalog items."""
    wb = load_workbook(filename=io.BytesIO(file_contents), data_only=True)
    ws = wb.active

    inserted_count = 0
    updated_count = 0
    errors = []

    rows = list(ws.iter_rows(values_only=True))
    if not rows or len(rows) < 2:
        return 0, 0, ["Excel spreadsheet is empty or missing headers"]

    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(v is None for v in row):
            continue

        try:
            category = str(row[1]).strip() if len(row) > 1 and row[1] else "General"
            subcategory = str(row[2]).strip() if len(row) > 2 and row[2] else "Standard"
            name = str(row[3]).strip() if len(row) > 3 and row[3] else "New Service"
            base_price = float(row[4]) if len(row) > 4 and row[4] is not None else 50.0
            max_demand = float(row[5]) if len(row) > 5 and row[5] is not None else 0.0
            max_discount = float(row[6]) if len(row) > 6 and row[6] is not None else 0.0
            is_active_val = str(row[7]).strip().lower() == "yes" if len(row) > 7 and row[7] is not None else True

            # Check if updating existing service by ID in column 0
            service = None
            if len(row) > 0 and row[0]:
                try:
                    s_uuid = uuid.UUID(str(row[0]).strip())
                    service = service_repository.get_service_by_id(db, s_uuid)
                except ValueError:
                    pass

            if service:
                service_repository.update_service(
                    db, service, name=name, category=category, subcategory=subcategory,
                    base_price=base_price, max_demand_increase=max_demand, max_discount=max_discount, is_active=is_active_val
                )
                updated_count += 1
            else:
                service_repository.create_service(
                    db, category=category, subcategory=subcategory, name=name,
                    base_price=base_price, max_demand_increase=max_demand, max_discount=max_discount, is_active=is_active_val
                )
                inserted_count += 1

        except Exception as e:
            errors.append(f"Row {idx}: {str(e)}")

    audit_repository.create_audit_log(
        db, actor_email=actor_email, actor_role="admin",
        action=f"Bulk Excel Catalog Import: {inserted_count} inserted, {updated_count} updated",
        risk_level="Info"
    )

    return inserted_count, updated_count, errors
