import os
import sys
import json
import hashlib
import datetime
from urllib.parse import urlparse, unquote
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_cat6_final_backups():
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    db_url = os.getenv("DATABASE_URL", "postgresql://postgres@localhost:5432/smartserve")
    p = urlparse(db_url)
    
    conn = psycopg2.connect(
        dbname=p.path.lstrip('/'),
        user=p.username,
        password=unquote(p.password or ''),
        host=p.hostname or 'localhost',
        port=p.port or 5432
    )
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT id, category, subcategory, name, base_price, max_demand_increase, max_discount,
               is_active, distinct_features, suggested_addons, created_at, updated_at
        FROM services
        WHERE category = '2. Cleaning & Pest Control'
        ORDER BY subcategory, name;
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    print(f"Fetched {len(rows)} Category 2 services from PostgreSQL.")
    
    backup_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backups"))
    os.makedirs(backup_dir, exist_ok=True)
    
    # 1. FINAL JSON
    json_path = os.path.join(backup_dir, "category2_cleaning_home_FINAL.json")
    final_records = []
    for r in rows:
        sa = r["suggested_addons"] or []
        real_addons = [a for a in sa if isinstance(a, dict) and not a.get("type") and a.get("name")]
        typed_blocks = [a for a in sa if isinstance(a, dict) and a.get("type")]
        
        final_records.append({
            "id": str(r["id"]),
            "category": r["category"],
            "subcategory": r["subcategory"],
            "name": r["name"],
            "base_price": float(r["base_price"]),
            "max_demand_increase": float(r["max_demand_increase"]) if r["max_demand_increase"] is not None else None,
            "max_discount": float(r["max_discount"]) if r["max_discount"] is not None else None,
            "is_active": r["is_active"],
            "distinct_features": r["distinct_features"] or [],
            "suggested_addons": sa,
            "real_addons_count": len(real_addons),
            "real_addons": real_addons,
            "typed_blocks_count": len(typed_blocks),
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None
        })
        
    final_json_doc = {
        "metadata": {
            "category": "2. Cleaning & Pest Control",
            "backup_type": "CATEGORY_FINAL",
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "total_services": len(final_records),
            "subcategories": {
                "Security Systems": len([r for r in final_records if r["subcategory"] == "Security Systems"])
            },
            "status": "PROTECTED_AND_PERSISTED"
        },
        "services": final_records
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(final_json_doc, f, indent=2, ensure_ascii=False)
    print(f"Saved FINAL JSON: {json_path} ({os.path.getsize(json_path)} bytes)")
    
    # 2. FINAL XLSX
    xlsx_path = os.path.join(backup_dir, "category2_cleaning_home_FINAL.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    ws_master = wb.create_sheet(title="ALL_SERVICES_FINAL")
    ws_master.views.sheetView[0].showGridLines = True
    
    columns = [
        "service_id", "service_name", "category", "subcategory", "base_price", "is_active",
        "distinct_features_json", "suggested_addons_json", "created_at", "updated_at"
    ]
    ws_master.append(columns)
    for col_idx in range(1, len(columns) + 1):
        c = ws_master.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        
    for r in final_records:
        row_data = [
            r["id"], r["name"], r["category"], r["subcategory"], r["base_price"], "TRUE" if r["is_active"] else "FALSE",
            json.dumps(r["distinct_features"], ensure_ascii=False),
            json.dumps(r["suggested_addons"], ensure_ascii=False),
            r["created_at"], r["updated_at"]
        ]
        ws_master.append(row_data)
        
    for col_letter in ['A', 'B', 'D', 'G', 'H']:
        ws_master.column_dimensions[col_letter].width = 40
        
    wb.save(xlsx_path)
    print(f"Saved FINAL XLSX: {xlsx_path} ({os.path.getsize(xlsx_path)} bytes)")

    # SQL Insert logic would go here if needed, but JSON/XLSX is standard.
    
if __name__ == "__main__":
    create_cat6_final_backups()
