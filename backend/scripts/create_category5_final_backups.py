import os
import sys
import json
import hashlib
import datetime
from urllib.parse import urlparse, unquote
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import RealDictCursor
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_cat5_final_backups():
    load_dotenv(os.path.join(os.path.dirname(__file__), "..", ".env"))
    db_url = os.getenv("DATABASE_URL", "postgresql://postgres@localhost:5432/smartserve")
    p = urlparse(db_url)
    
    conn = psycopg2.connect(
        dbname=p.path.lstrip('/'),
        user=p.username,
        password=unquote(p.password or ''),
        host=p.hostname or 'localhost',
        port=p.port or 5432
    )
    cur = conn.cursor(cursor_factory=RealDictCursor)
    
    cur.execute("""
        SELECT id, category, subcategory, name, base_price, max_demand_increase, max_discount,
               is_active, distinct_features, suggested_addons, created_at, updated_at
        FROM services
        WHERE category = '5. Electrician, Plumber, Carpenter & Home Repairs'
        ORDER BY subcategory, name;
    """)
    rows = cur.fetchall()
    cur.close()
    conn.close()
    
    print(f"Fetched {len(rows)} Category 5 services from PostgreSQL.")
    assert len(rows) == 39, f"Expected 39 services, got {len(rows)}"
    
    backup_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "backups"))
    os.makedirs(backup_dir, exist_ok=True)
    
    # 1. FINAL JSON
    json_path = os.path.join(backup_dir, "category5_electrician_plumber_carpenter_home_repairs_FINAL.json")
    final_records = []
    for r in rows:
        sa = r["suggested_addons"] or []
        real_addons = [a for a in sa if isinstance(a, dict) and not a.get("type") and a.get("name")]
        typed_blocks = [a for a in sa if isinstance(a, dict) and a.get("type")]
        
        final_records.append({
            "id": str(r["id"]),
            "category": r["category"],
            "subcategory": r["subcategory"],
            "name": r["name"],
            "base_price": float(r["base_price"]),
            "max_demand_increase": float(r["max_demand_increase"]) if r["max_demand_increase"] is not None else None,
            "max_discount": float(r["max_discount"]) if r["max_discount"] is not None else None,
            "is_active": r["is_active"],
            "distinct_features": r["distinct_features"] or [],
            "suggested_addons": sa,
            "real_addons_count": len(real_addons),
            "real_addons": real_addons,
            "typed_blocks_count": len(typed_blocks),
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None
        })
        
    final_json_doc = {
        "metadata": {
            "category": "5. Electrician, Plumber, Carpenter & Home Repairs",
            "backup_type": "CATEGORY_FINAL",
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "total_services": len(final_records),
            "subcategories": {
                "Carpenter": len([r for r in final_records if r["subcategory"] == "Carpenter"]),
                "Electrician": len([r for r in final_records if r["subcategory"] == "Electrician"]),
                "Plumber": len([r for r in final_records if r["subcategory"] == "Plumber"])
            },
            "status": "PROTECTED_AND_PERSISTED"
        },
        "services": final_records
    }
    
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(final_json_doc, f, indent=2, ensure_ascii=False)
    print(f"Saved FINAL JSON: {json_path} ({os.path.getsize(json_path)} bytes)")
    
    # 2. FINAL XLSX
    xlsx_path = os.path.join(backup_dir, "category5_electrician_plumber_carpenter_home_repairs_FINAL.xlsx")
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    navy_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
    teal_fill = PatternFill(start_color="0D9488", end_color="0D9488", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell_font = Font(name="Calibri", size=10)
    border_thin = Side(border_style="thin", color="CBD5E1")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    master_cols = [
        "service_id", "service_name", "category", "subcategory", "base_price", "is_active",
        "description", "highlights", "included", "excluded", "process_steps",
        "tools_materials", "customer_setup", "aftercare", "expected_results",
        "important_notes", "warranty", "faqs", "dos", "donts", "tips", "existing_add_ons"
    ]
    
    # Master sheet
    ws_all = wb.create_sheet(title="MASTER_CATEGORY_5")
    ws_all.views.sheetView[0].showGridLines = True
    ws_all.append(master_cols)
    ws_all.row_dimensions[1].height = 25
    for col_idx in range(1, len(master_cols) + 1):
        c = ws_all.cell(row=1, column=col_idx)
        c.font = header_font
        c.fill = navy_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = cell_border
        
    def extract_row_values(r):
        sa = r["suggested_addons"]
        desc_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "description"), None)
        hl_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "highlights"), None)
        exc_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") in ["excluded_scope", "exclusions"]), None)
        proc_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "process_steps"), None)
        ac_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "aftercare_precautions"), None)
        tm_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "tools_materials"), None)
        cs_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "customer_setup"), None)
        er_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "expected_results"), None)
        note_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "important_notes"), None)
        warr_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "warranty"), None)
        faq_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "faqs"), None)
        tip_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "tips"), None)
        dd_obj = next((a for a in sa if isinstance(a, dict) and a.get("type") == "dos_donts"), None)
        
        desc = desc_obj.get("text") if desc_obj else ""
        hl = "\n".join(f"• {h}" for h in hl_obj.get("items", [])) if hl_obj else ""
        inc = "\n".join(f"• {i}" for i in r["distinct_features"])
        exc = "\n".join(f"• {e}" for e in exc_obj.get("items", [])) if exc_obj else ""
        proc = "\n\n".join(f"Step {p['step_number']}: {p['title']}\n{p['description']}" for p in proc_obj.get("steps", [])) if proc_obj else ""
        tm = ", ".join(tm_obj.get("tools", [])) if tm_obj else ""
        cs = "\n".join(f"• {c}" for c in cs_obj.get("requirements", [])) if cs_obj else ""
        ac = "\n".join(f"• {a}" for a in ac_obj.get("aftercare", [])) if ac_obj else ""
        er = "\n".join(f"• {e}" for e in er_obj.get("items", [])) if er_obj else ""
        nt = "\n".join(f"• {n}" for n in note_obj.get("items", [])) if note_obj else ""
        warr = warr_obj.get("details", "") if warr_obj else "N/A"
        faq = "\n\n".join(f"Q: {f['question']}\nA: {f['answer']}" for f in faq_obj.get("items", [])) if faq_obj else ""
        dos = "\n".join(f"• {d}" for d in dd_obj.get("dos", [])) if dd_obj else ""
        donts = "\n".join(f"• {dt}" for dt in dd_obj.get("donts", [])) if dd_obj else ""
        tips = "\n".join(f"• {t}" for t in tip_obj.get("items", [])) if tip_obj else ""
        addons = "\n".join(f"• {a.get('name')}: Rs.{a.get('price')}" for a in r["real_addons"]) if r["real_addons"] else "None"
        
        return [
            r["id"], r["name"], r["category"], r["subcategory"], r["base_price"], "True",
            desc, hl, inc, exc, proc, tm, cs, ac, er, nt, warr, faq, dos, donts, tips, addons
        ]

    for r in final_records:
        row_vals = extract_row_values(r)
        ws_all.append(row_vals)
        curr_row = ws_all.max_row
        ws_all.row_dimensions[curr_row].height = 65
        for col_idx in range(1, len(row_vals) + 1):
            c = ws_all.cell(row=curr_row, column=col_idx)
            c.font = cell_font
            c.border = cell_border
            c.alignment = Alignment(vertical="top", wrap_text=True)
            
    for col_idx in range(1, len(master_cols) + 1):
        col_letter = get_column_letter(col_idx)
        if col_idx in [1, 2, 4]:
            ws_all.column_dimensions[col_letter].width = 30
        elif col_idx in [3, 5, 6]:
            ws_all.column_dimensions[col_letter].width = 16
        else:
            ws_all.column_dimensions[col_letter].width = 45
            
    # Subcategory individual sheets
    for sc in ["Carpenter", "Electrician", "Plumber"]:
        ws_sc = wb.create_sheet(title=sc.upper())
        ws_sc.views.sheetView[0].showGridLines = True
        ws_sc.append(master_cols)
        ws_sc.row_dimensions[1].height = 25
        for col_idx in range(1, len(master_cols) + 1):
            c = ws_sc.cell(row=1, column=col_idx)
            c.font = header_font
            c.fill = teal_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = cell_border
            
        sc_records = [r for r in final_records if r["subcategory"] == sc]
        for r in sc_records:
            row_vals = extract_row_values(r)
            ws_sc.append(row_vals)
            curr_row = ws_sc.max_row
            ws_sc.row_dimensions[curr_row].height = 65
            for col_idx in range(1, len(row_vals) + 1):
                c = ws_sc.cell(row=curr_row, column=col_idx)
                c.font = cell_font
                c.border = cell_border
                c.alignment = Alignment(vertical="top", wrap_text=True)
                
        for col_idx in range(1, len(master_cols) + 1):
            col_letter = get_column_letter(col_idx)
            if col_idx in [1, 2, 4]:
                ws_sc.column_dimensions[col_letter].width = 30
            elif col_idx in [3, 5, 6]:
                ws_sc.column_dimensions[col_letter].width = 16
            else:
                ws_sc.column_dimensions[col_letter].width = 45
                
    wb.save(xlsx_path)
    print(f"Saved FINAL XLSX: {xlsx_path} ({os.path.getsize(xlsx_path)} bytes)")
    
    # 3. FINAL SQL (Restorable)
    sql_path = os.path.join(backup_dir, "category5_electrician_plumber_carpenter_home_repairs_FINAL.sql")
    with open(sql_path, "w", encoding="utf-8") as f:
        f.write("-- SmartServe Restorable Category 5 Backup\n")
        f.write(f"-- Timestamp: {datetime.datetime.now(datetime.timezone.utc).isoformat()}\n")
        f.write(f"-- Target: Category 5 - Electrician, Plumber, Carpenter & Home Repairs (39 Services)\n\n")
        f.write("BEGIN;\n\n")
        
        for r in final_records:
            sid = r["id"]
            name = r["name"].replace("'", "''")
            cat = r["category"].replace("'", "''")
            subcat = r["subcategory"].replace("'", "''")
            price = r["base_price"]
            df_json = json.dumps(r["distinct_features"]).replace("'", "''")
            sa_json = json.dumps(r["suggested_addons"]).replace("'", "''")
            
            f.write(f"""INSERT INTO services (id, name, category, subcategory, base_price, is_active, distinct_features, suggested_addons, updated_at)
VALUES ('{sid}', '{name}', '{cat}', '{subcat}', {price}, true, '{df_json}'::jsonb, '{sa_json}'::jsonb, NOW())
ON CONFLICT (id) DO UPDATE SET
    name = EXCLUDED.name,
    category = EXCLUDED.category,
    subcategory = EXCLUDED.subcategory,
    base_price = EXCLUDED.base_price,
    is_active = EXCLUDED.is_active,
    distinct_features = EXCLUDED.distinct_features,
    suggested_addons = EXCLUDED.suggested_addons,
    updated_at = NOW();\n\n""")
            
        f.write("COMMIT;\n")
        
    print(f"Saved FINAL SQL:  {sql_path} ({os.path.getsize(sql_path)} bytes)")
    
    # Calculate SHA256 hashes
    def get_sha256(filepath):
        h = hashlib.sha256()
        with open(filepath, "rb") as f:
            while chunk := f.read(65536):
                h.update(chunk)
        return h.hexdigest()
        
    j_sha = get_sha256(json_path)
    x_sha = get_sha256(xlsx_path)
    s_sha = get_sha256(sql_path)
    
    print("\n" + "=" * 80)
    print("PHASE 12 PERMANENT CATEGORY 5 BACKUPS COMPLETE & HASHED:")
    print(f"  JSON SHA-256: {j_sha} | {json_path}")
    print(f"  XLSX SHA-256: {x_sha} | {xlsx_path}")
    print(f"  SQL  SHA-256: {s_sha} | {sql_path}")
    print("=" * 80)
    
    return json_path, xlsx_path, sql_path, j_sha, x_sha, s_sha

if __name__ == "__main__":
    create_cat5_final_backups()
